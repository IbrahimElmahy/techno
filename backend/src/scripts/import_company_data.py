"""Import the company's real master data from the client's Excel workbook (v4).

Source: "داتا الاصناف.xlsx"
- Sheet «الاصناف و الفئات»: 7 side-by-side category blocks, each
  (الصنف | العدد | سعر اللسته | نسبه الخصم | قيمه النقطه). A row that has a name but no quantity
  and no price is an INLINE sub-category header for the rows beneath it.
- Sheet «المناديب»: reps + their cash boxes, the warehouse list, and customer account balances
  (مدين; may be negative).

Maps to: Item (name/category/sale_price/default_discount_pct/point value), opening stock in the main
warehouse, Warehouse, sales-rep users, Customer + opening receivable balances.

Idempotent: existing items/warehouses/customers are matched by name and skipped, so re-running only
adds what is missing. Run:  python -m src.scripts.import_company_data <path-to-xlsx>
"""
from __future__ import annotations

import sys
from decimal import Decimal, InvalidOperation

from sqlalchemy import select
from sqlalchemy.orm import Session

from src.core.security import hash_password
from src.models.catalog import Item, ItemKind
from src.models.customer import Customer, CustomerAccount
from src.models.ledger import Direction
from src.models.loyalty import ProductPointValue
from src.models.org import Territory
from src.models.role import Role, RoleName
from src.models.stock import LocationKind, StockDirection
from src.models.user import User
from src.models.warehouse import Warehouse, WarehouseType
from src.services import (
    account_resolver,
    customer_service,
    ledger_service,
    lookup_service,
    stock_service,
)
from src.services.ledger_service import LineInput

BLOCK_STARTS = [0, 4, 10, 16, 22, 28, 34]
ITEMS_SHEET = "الاصناف و الفئات"
REPS_SHEET = "المناديب"
MAIN_WAREHOUSE = "المخزن الرئيسى"
DEFAULT_UOM = "قطعة"


def _num(v) -> Decimal | None:
    """Excel cells arrive as int/float/str; return a Decimal or None when not a number."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return None


def parse_items(ws) -> list[dict]:
    """Flatten the side-by-side category blocks into item dicts."""
    rows = list(ws.iter_rows(values_only=True))

    def cell(r, c):
        return r[c] if c < len(r) else None

    items: list[dict] = []
    for start in BLOCK_STARTS:
        category = (str(rows[0][start]).strip() if rows[0][start] else "")
        for r in rows[2:]:
            raw_name = cell(r, start)
            if raw_name is None or not str(raw_name).strip():
                continue
            name = str(raw_name).strip()
            qty, price = cell(r, start + 1), cell(r, start + 2)
            # name only, no numbers -> inline sub-category header for the rows below
            if _num(qty) is None and _num(price) is None:
                category = name
                continue
            if name == "0":  # junk row in the source file
                continue
            items.append({
                "name": name,
                "category": category,
                "quantity": _num(qty) or Decimal("0"),
                "price": _num(price),
                # source stores a fraction (0.1) — the system stores a percentage (10.00)
                "discount_pct": (_num(cell(r, start + 3)) or Decimal("0")) * 100,
                "points": _num(cell(r, start + 4)) or Decimal("0"),
            })
    return items


def parse_reps_sheet(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))

    def cell(r, c):
        return r[c] if c < len(r) else None

    warehouses = [str(cell(r, 8)).strip() for r in rows[1:8]
                  if cell(r, 8) and str(cell(r, 8)).strip() != "المخازن"]
    reps, boxes = [], []
    for col in (0, 4):
        for r in rows[1:5]:
            if cell(r, col):
                reps.append(str(cell(r, col)).strip())
            if cell(r, col + 1):
                boxes.append(str(cell(r, col + 1)).strip())
    accounts = []
    for col in (0, 4):
        for r in rows[6:]:
            name, bal = cell(r, col), cell(r, col + 1)
            if name and str(name).strip() and str(name).strip() != "الحساب":
                accounts.append({"name": str(name).strip(), "balance": _num(bal) or Decimal("0")})
    return {"warehouses": warehouses, "reps": reps, "boxes": boxes, "accounts": accounts}


def import_workbook(db: Session, path: str) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    items = parse_items(wb[ITEMS_SHEET])
    meta = parse_reps_sheet(wb[REPS_SHEET]) if REPS_SHEET in wb.sheetnames else {
        "warehouses": [], "reps": [], "boxes": [], "accounts": []}

    admin = db.scalar(select(User).where(User.username == "admin"))
    actor = admin.id
    summary = {"items_created": 0, "items_skipped": 0, "no_price": 0, "stock_posted": 0,
               "warehouses_created": 0, "reps_created": 0, "customers_created": 0,
               "customers_skipped": 0, "opening_balance_total": "0", "categories": 0}

    # --- Warehouses (from the sheet; main one first) ---
    wh_names = meta["warehouses"] or [MAIN_WAREHOUSE]
    wh_by_name: dict[str, Warehouse] = {}
    for name in wh_names:
        wh = db.scalar(select(Warehouse).where(Warehouse.name == name))
        if wh is None:
            wh = Warehouse(name=name, warehouse_type=WarehouseType.central)
            db.add(wh)
            db.flush()
            summary["warehouses_created"] += 1
        wh_by_name[name] = wh
    main_wh = wh_by_name.get(MAIN_WAREHOUSE) or next(iter(wh_by_name.values()))

    # --- Item categories -> the configurable `item_category` lookup ---
    categories = sorted({i["category"] for i in items if i["category"]})
    existing_cats = {o.value for o in lookup_service.list_options(db, "item_category")}
    for cat in categories:
        if cat not in existing_cats:
            try:
                lookup_service.create_option(db, category="item_category", value=cat, label=cat)
                summary["categories"] += 1
            except Exception:  # pragma: no cover - duplicate/racy, harmless
                pass

    # --- Items (products) + point values + opening stock ---
    existing_items = {i.name: i for i in db.scalars(select(Item)).all()}
    n_products = db.query(Item).filter(Item.kind == ItemKind.product).count()
    for row in items:
        if row["name"] in existing_items:
            summary["items_skipped"] += 1
            continue
        n_products += 1
        if row["price"] is None:
            summary["no_price"] += 1
        item = Item(
            code=f"PR-{n_products:06d}", name=row["name"], kind=ItemKind.product,
            unit_of_measure=DEFAULT_UOM, sale_price=row["price"], category=row["category"] or None,
            default_discount_pct=row["discount_pct"], default_warehouse_id=main_wh.id,
        )
        db.add(item)
        db.flush()
        existing_items[item.name] = item
        summary["items_created"] += 1

        if row["points"] and row["points"] > 0:  # fractional point values (v4)
            db.add(ProductPointValue(item_id=item.id, point_value=row["points"], updated_by=actor))

        if row["quantity"] and row["quantity"] > 0:  # opening stock into the main warehouse
            stock_service.post_movement(
                db, item_id=item.id, location_kind=LocationKind.warehouse, location_id=main_wh.id,
                movement_type="opening_stock", direction=StockDirection.in_,
                quantity=row["quantity"], actor_user_id=actor,
                source_doc_type="import", source_doc_id=item.id,
            )
            summary["stock_posted"] += 1
    db.flush()

    # --- Sales reps from the sheet ---
    territory = db.scalar(select(Territory))
    rep_role = db.scalar(select(Role).where(Role.name == RoleName.sales_rep))
    if rep_role is None:
        rep_role = Role(name=RoleName.sales_rep)
        db.add(rep_role)
        db.flush()
    rep_user = None
    for idx, rep_name in enumerate(meta["reps"], 1):
        existing = db.scalar(select(User).where(User.full_name == rep_name))
        if existing is not None:
            rep_user = rep_user or existing
            continue
        u = User(username=f"rep_import_{idx}", password_hash=hash_password("rep123456"),
                 role_id=rep_role.id, full_name=rep_name,
                 branch_id=territory.branch_id if territory else None,
                 territory_id=territory.id if territory else None)
        db.add(u)
        db.flush()
        rep_user = rep_user or u
        summary["reps_created"] += 1
    if rep_user is None:
        rep_user = db.scalar(select(User).join(Role).where(Role.name == RoleName.sales_rep)) or admin

    # --- Customers + opening receivable balances ---
    existing_customers = {c.name for c in db.scalars(select(Customer)).all()}
    total_open = Decimal("0")
    opening_lines: list[LineInput] = []
    for acc in meta["accounts"]:
        if acc["name"] in existing_customers:
            summary["customers_skipped"] += 1
            continue
        result = customer_service.create_customer(
            db, name=acc["name"], customer_type="trader", rep_id=rep_user.id,
            territory_id=territory.id if territory else None, phone=None, actor_user_id=actor)
        cust = result.customer
        existing_customers.add(cust.name)
        summary["customers_created"] += 1
        bal = acc["balance"]
        if bal and bal != 0:
            # مدين (positive) = the customer owes us -> debit their receivable account.
            ca = db.scalar(select(CustomerAccount).where(CustomerAccount.customer_id == cust.id))
            direction = Direction.debit if bal > 0 else Direction.credit
            opening_lines.append(LineInput(ca.account_id, direction, abs(bal)))
            total_open += bal

    if opening_lines:
        # Balance the customer receivables against opening equity in ONE entry.
        equity = account_resolver.opening_balance_equity_account(db)
        net = sum((ln.amount if ln.direction == Direction.debit else -ln.amount)
                  for ln in opening_lines)
        if net > 0:
            opening_lines.append(LineInput(equity.id, Direction.credit, net))
        elif net < 0:
            opening_lines.append(LineInput(equity.id, Direction.debit, -net))
        ledger_service.post_entry(
            db, entry_type="opening_balance", actor_user_id=actor, lines=opening_lines,
            description="أرصدة افتتاحية للعملاء (استيراد)")
    summary["opening_balance_total"] = str(total_open)

    db.commit()
    return summary


def main() -> None:
    from src.core.db import SessionLocal

    path = sys.argv[1] if len(sys.argv) > 1 else "داتا الاصناف.xlsx"
    db = SessionLocal()
    try:
        print(import_workbook(db, path))
    finally:
        db.close()


if __name__ == "__main__":
    main()
