"""Catalog router (T008). FR-001–005. System-generated editable code; kind/price validation."""
from __future__ import annotations

from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, Query, status, UploadFile
from pydantic import BaseModel
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from src.auth.dependencies import CurrentUser, require_capability
from src.auth.rbac import (
    CAP_CATALOG_READ,
    CAP_CATALOG_WRITE,
    CAP_PRODUCT_POINTS_WRITE,
    CAP_PURCHASE_WRITE,
    CAP_STOCK_READ,
    role_has_capability,
)
from src.core.db import get_db
from src.core.money import to_money, to_qty
from src.models.catalog import (
    Item,
    ItemKind,
    ItemPrice,
    ItemSerial,
    ItemUnit,
    PriceTier,
    SerialStatus,
)
from src.models.stock import LocationKind
from src.lib import item_card as item_card_lib
from src.services import audit_service, item_profile_service, serial_service
from src.services.serial_service import SerialError

router = APIRouter(tags=["catalog"], prefix="/items")


class ItemCreate(BaseModel):
    name: str
    kind: ItemKind
    unit_of_measure: str
    purchase_price: Decimal | None = None
    sale_price: Decimal | None = None
    is_serialized: bool = False
    default_warehouse_id: int | None = None
    category: str | None = None
    default_discount_pct: Decimal | None = None
    # (011) advisory planning thresholds + expiry-batch tracking
    min_stock: Decimal | None = None
    max_stock: Decimal | None = None
    is_perishable: bool = False
    # التعبئة + a free note (a5 parity).
    piece_name: str | None = None
    pieces_per_unit: Decimal | None = None
    description: str | None = None
    # (031) The three things an item carries that do not live on its own row. They used to be
    # three more HTTP calls the screen made AFTER the item existed, which meant a failure on any
    # of them left an item created and half-configured while the screen said «اتسجّل الصنف».
    # Taken here, they are written inside the SAME transaction: the item and everything about it
    # either exist together or not at all.
    #
    # Typed as `list[dict] | None` at this point in the file because `TierPrice` and `UnitIn` are
    # declared below; `_apply_*` validates them through those models before anything is written.
    tiers: list[dict] | None = None
    units: list[dict] | None = None
    point_value: Decimal | None = None


class ItemUpdate(BaseModel):
    code: str | None = None
    name: str | None = None
    purchase_price: Decimal | None = None
    sale_price: Decimal | None = None
    is_serialized: bool | None = None
    active: bool | None = None
    default_warehouse_id: int | None = None
    category: str | None = None
    default_discount_pct: Decimal | None = None
    min_stock: Decimal | None = None          # (011)
    max_stock: Decimal | None = None
    is_perishable: bool | None = None
    piece_name: str | None = None
    pieces_per_unit: Decimal | None = None
    description: str | None = None


class ItemOut(BaseModel):
    id: int
    code: str
    name: str
    kind: ItemKind
    unit_of_measure: str
    purchase_price: Decimal | None
    sale_price: Decimal | None
    is_serialized: bool
    active: bool
    default_warehouse_id: int | None = None
    category: str | None = None
    default_discount_pct: Decimal | None = None
    # (011) advisory planning thresholds + expiry-batch tracking
    min_stock: Decimal | None = None
    max_stock: Decimal | None = None
    is_perishable: bool = False
    # التعبئة + a free note (a5 parity).
    piece_name: str | None = None
    pieces_per_unit: Decimal | None = None
    description: str | None = None
    # Total on-hand across all locations — filled on the list endpoint (one grouped query).
    on_hand: Decimal | None = None
    # Their list carries the مستهلك price as a column; it lives in its own table, so the list
    # endpoint fills it in bulk rather than making the screen ask per row.
    consumer_price: Decimal | None = None


def _out(it: Item) -> ItemOut:
    return ItemOut(
        id=it.id, code=it.code, name=it.name, kind=it.kind,
        unit_of_measure=it.unit_of_measure, purchase_price=it.purchase_price,
        sale_price=it.sale_price, is_serialized=it.is_serialized, active=it.active,
        default_warehouse_id=it.default_warehouse_id, category=it.category,
        default_discount_pct=it.default_discount_pct,
        min_stock=it.min_stock, max_stock=it.max_stock, is_perishable=it.is_perishable,
        piece_name=it.piece_name, pieces_per_unit=it.pieces_per_unit,
        description=it.description,
    )


def _next_code(db: Session, kind: ItemKind) -> str:
    prefix = "RM" if kind == ItemKind.raw_material else "PR"
    n = db.scalar(select(func.count()).select_from(Item).where(Item.kind == kind)) or 0
    return f"{prefix}-{n + 1:06d}"


@router.get("", response_model=list[ItemOut])
def list_items(
    kind: ItemKind | None = None,
    category: str | None = None,
    q: str | None = None,
    active: bool | None = None,
    warehouse_id: int | None = None,
    stock_filter: str | None = None,  # all | in_stock | out_of_stock | negative | moved
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_READ)),
    db: Session = Depends(get_db),
) -> list[ItemOut]:
    """List items with search + filters; each row carries its total on-hand quantity.

    On-hand comes from ONE grouped query over the movements, so filtering by stock costs the
    same as listing.
    """
    stmt = item_profile_service.apply_filters(
        select(Item), q=q, kind=kind.value if kind else None, category=category,
        active=active, warehouse_id=warehouse_id,
    )
    rows = list(db.scalars(stmt).all())
    on_hand = item_profile_service.bulk_on_hand(db, [i.id for i in rows])
    moved = (item_profile_service.bulk_has_movement(db, [i.id for i in rows])
             if stock_filter == "moved" else None)
    rows = item_profile_service.filter_by_stock(rows, on_hand, stock_filter, moved)
    # Looked up after the stock filter, so the extra two queries only cover rows that survive it.
    ids = [i.id for i in rows]
    consumer = item_profile_service.bulk_tier_price(db, ids, PriceTier.consumer)
    out = []
    for i in rows:
        o = _out(i)
        o.on_hand = on_hand.get(i.id, Decimal("0.000"))
        o.consumer_price = consumer.get(i.id)
        out.append(o)
    return out


def _apply_tiers(db: Session, item: Item, tiers: list["TierPrice"], *, actor_user_id: int) -> None:
    """Write the sale-price tiers. Shared by `POST /items` and `PUT /items/{id}/prices` so the
    validation and the price-change log cannot come to differ between creating and editing."""
    if item.kind != ItemKind.product:
        raise HTTPException(422, {"code": "validation", "message": "only products have sale prices"})
    for tp in tiers:
        if tp.price < 0:
            raise HTTPException(422, {"code": "validation", "message": "price must be ≥ 0"})
    for tp in tiers:
        row = db.scalar(
            select(ItemPrice).where(ItemPrice.item_id == item.id, ItemPrice.tier == tp.tier)
        )
        # Log the move BEFORE writing it, so the history keeps the previous price (027).
        item_profile_service.record_price_change(
            db, item_id=item.id, field_name=tp.tier.value,
            old_value=row.price if row is not None else None,
            new_value=to_money(tp.price), actor_user_id=actor_user_id)
        if row is None:
            db.add(ItemPrice(item_id=item.id, tier=tp.tier, price=to_money(tp.price),
                             discount_pct=tp.discount_pct or 0, vat_pct=tp.vat_pct or 0))
        else:
            row.price = to_money(tp.price)
    db.flush()


def _apply_units(db: Session, item: Item, units: list["UnitIn"]) -> None:
    """Replace the whole alternate-unit set — which is what makes removing one possible."""
    seen = {item.unit_of_measure}
    for u in units:
        if u.factor <= 0:
            raise HTTPException(422, {"code": "validation", "message": "factor must be > 0"})
        if u.name in seen:
            raise HTTPException(422, {"code": "validation",
                                      "message": f"duplicate unit name '{u.name}'"})
        seen.add(u.name)
    db.execute(delete(ItemUnit).where(ItemUnit.item_id == item.id))
    for u in units:
        db.add(ItemUnit(item_id=item.id, name=u.name, factor=to_qty(u.factor)))
    db.flush()


def _apply_point_value(db: Session, item: Item, value, *, actor_user_id: int) -> None:
    """Loyalty points per piece. Products only — a raw material has none to give."""
    from src.models.loyalty import ProductPointValue

    if item.kind != ItemKind.product:
        raise HTTPException(422, {"code": "validation",
                                  "message": "Point values apply to products only"})
    if value < 0:
        raise HTTPException(422, {"code": "validation", "message": "point_value must be ≥ 0"})
    ppv = db.scalar(select(ProductPointValue).where(ProductPointValue.item_id == item.id))
    if ppv is None:
        db.add(ProductPointValue(item_id=item.id, point_value=value, updated_by=actor_user_id))
    else:
        ppv.point_value = value
        ppv.updated_by = actor_user_id
    db.flush()




# --- استيراد الأصناف من إكسل -------------------------------------------------------

_IMPORT_HEADERS = [
    "الاسم", "الفئة", "الوحدة",
    "سعر الجملة", "سعر التجاري", "سعر نص جملة", "سعر نص تجاري",
    "سعر المستهلك", "سعر اللستة",
    "أقل مخزون", "أقصى مخزون", "ملاحظات",
]

_TIER_COLUMNS: list[tuple[str, PriceTier]] = [
    ("سعر الجملة", PriceTier.wholesale),
    ("سعر التجاري", PriceTier.commercial),
    ("سعر نص جملة", PriceTier.semi_wholesale),
    ("سعر نص تجاري", PriceTier.semi_commercial),
    ("سعر المستهلك", PriceTier.consumer),
    ("سعر اللستة", PriceTier.list_price),
]


@router.get("/import-template")
def items_import_template(
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_READ)),
):
    """قالب الاستيراد — نفس الأعمدة اللي المستورد بيقراها، بسطرين مثال."""
    import io as _io

    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الأصناف"
    ws.append(_IMPORT_HEADERS)
    ws.append(["سخام 4 بوصة", "مواسير", "قطعة", 120, 135, 110, 125, 150, None, None, None, None])
    ws.append(["صنف تاني بدون أسعار اختيارية", "عدد وأدوات", "قطعة"])
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=items-import-template.xlsx"},
    )


def _dec(v) -> Decimal | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        d = Decimal(str(v).strip())
        return d
    except Exception:
        return None


@router.post("/import-excel")
async def import_items_excel(
    file: UploadFile = File(...),
    current: CurrentUser = Depends(require_capability(CAP_CATALOG_WRITE)),
    db: Session = Depends(get_db),
) -> dict:
    """استيراد أصناف من ملف إكسل بأعمدة القالب.

    كل صف بيتحفظ في معاملة مستقلة (savepoint): الصنفي اللي بيغلطوا مايوقفوش الباقي،
    واللي اتكرر بالاسم بيتعّد «تم تخطيه» بدل ما يتكرر.
    """
    import io as _io

    import openpyxl

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, {
            "code": "bad_file",
            "message": "الملف مش إكسل مقروء — نزّل القالب واملأ بنفس الأعمدة.",
        })
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            {"code": "empty", "message": "الملف فاضي."})
    header = [(str(cell).strip() if cell is not None else "") for cell in rows[0]]
    if "الاسم" not in header:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, {
            "code": "bad_headers",
            "message": "أول صف لازم يكون عناوين الأعمدة ويسبقها عمود «الاسم» — نزّل القالب.",
        })

    def col(title: str) -> int:
        return header.index(title)

    i_name = col("الاسم")
    i_cat = header.index("الفئة") if "الفئة" in header else -1
    i_unit = header.index("الوحدة") if "الوحدة" in header else -1
    i_min = header.index("أقل مخزون") if "أقل مخزون" in header else -1
    i_max = header.index("أقصى مخزون") if "أقصى مخزون" in header else -1
    i_note = header.index("ملاحظات") if "ملاحظات" in header else -1
    tier_cols = [(header.index(t), tier) for t, tier in _TIER_COLUMNS if t in header]
    i_whole = header.index("سعر الجملة") if "سعر الجملة" in header else -1

    created = skipped = failed = 0
    errors: list[dict] = []
    for idx, r in enumerate(rows[1:], start=2):
        def cell(i: int):
            return r[i] if 0 <= i < len(r) else None

        name = str(cell(i_name)).strip() if cell(i_name) is not None else ""
        if not name:
            continue
        try:
            existing = db.scalar(select(Item).where(func.lower(Item.name) == name.lower()))
            if existing is not None:
                skipped += 1
                continue
            unit = str(cell(i_unit)).strip() if cell(i_unit) is not None and str(cell(i_unit)).strip() else "قطعة"
            sale_price = _dec(cell(i_whole)) if i_whole >= 0 else None
            item = Item(
                code=_next_code(db, ItemKind.product), name=name, kind=ItemKind.product,
                unit_of_measure=unit, sale_price=sale_price,
                category=(str(cell(i_cat)).strip() if i_cat >= 0 and cell(i_cat) is not None else None),
                min_stock=_dec(cell(i_min)) if i_min >= 0 else None,
                max_stock=_dec(cell(i_max)) if i_max >= 0 else None,
                description=(str(cell(i_note)).strip() if i_note >= 0 and cell(i_note) is not None else None),
            )
            db.add(item)
            db.flush()
            tiers = []
            for ci, tier in tier_cols:
                v = _dec(cell(ci))
                if v is not None:
                    tiers.append(TierPrice(tier=tier, price=v))
            if item.sale_price is not None and not any(t.tier == PriceTier.wholesale for t in tiers):
                tiers.append(TierPrice(tier=PriceTier.wholesale, price=item.sale_price))
            if tiers:
                _apply_tiers(db, item, tiers, actor_user_id=current.id)
            db.commit()
            created += 1
        except HTTPException as exc:
            db.rollback()
            failed += 1
            msg = (exc.detail or {}).get("message") if isinstance(exc.detail, dict) else str(exc.detail)
            errors.append({"row": idx, "name": name, "message": msg or "بيانات غير صحيحة"})
        except Exception as exc:
            db.rollback()
            failed += 1
            errors.append({"row": idx, "name": name, "message": str(exc)})
        if len(errors) >= 50:
            errors.append({"row": 0, "name": "", "message": "… وفيه أخطاء تانية اتشالت من العرض"})
            break
    return {"created": created, "skipped": skipped, "failed": failed, "errors": errors}


@router.post("", response_model=ItemOut, status_code=status.HTTP_201_CREATED)
def create_item(
    body: ItemCreate,
    current: CurrentUser = Depends(require_capability(CAP_CATALOG_WRITE)),
    db: Session = Depends(get_db),
) -> ItemOut:
    if body.kind == ItemKind.raw_material and body.sale_price is not None:
        raise HTTPException(422, {"code": "validation", "message": "raw material has no sale price"})
    if body.kind == ItemKind.product and body.purchase_price is not None:
        raise HTTPException(422, {"code": "validation", "message": "product has no purchase price"})
    item = Item(
        code=_next_code(db, body.kind),
        name=body.name,
        kind=body.kind,
        unit_of_measure=body.unit_of_measure,
        purchase_price=body.purchase_price,
        sale_price=body.sale_price,
        is_serialized=body.is_serialized,
        default_warehouse_id=body.default_warehouse_id, category=body.category,
        default_discount_pct=body.default_discount_pct or 0,
        min_stock=body.min_stock, max_stock=body.max_stock,   # (011) advisory
        is_perishable=body.is_perishable,
        piece_name=body.piece_name, pieces_per_unit=body.pieces_per_unit,
        description=body.description,
    )
    db.add(item)
    db.flush()

    # Everything else the form sends, in this transaction. Each helper raises on bad input, and
    # because nothing has been committed yet the item goes with it.
    if body.tiers:
        _apply_tiers(db, item, [TierPrice(**t) for t in body.tiers], actor_user_id=current.id)
    if body.units:
        _apply_units(db, item, [UnitIn(**u) for u in body.units])
    if body.point_value is not None:
        # Checked here rather than trusted: point values are a different capability from the
        # catalogue, and a purchasing manager may create items without being allowed to price
        # loyalty. Refusing BEFORE the commit is what keeps «اتسجّل الصنف» honest — the older
        # shape created the item, got a 403 on a second call, and reported success anyway.
        if not role_has_capability(current.role, CAP_PRODUCT_POINTS_WRITE):
            raise HTTPException(403, {"code": "forbidden",
                                      "message": "لا تملك صلاحية تحديد نقاط المنتج"})
        _apply_point_value(db, item, body.point_value, actor_user_id=current.id)

    db.commit()
    return _out(item)


class TierPrice(BaseModel):
    tier: PriceTier
    price: Decimal
    # Each tier carries its own allowance: a wholesaler and a walk-in do not get the same one.
    discount_pct: Decimal = Decimal("0")
    vat_pct: Decimal = Decimal("0")


class ItemPricesOut(BaseModel):
    item_id: int
    base_sale_price: Decimal | None
    tiers: list[TierPrice]


class ItemPricesSet(BaseModel):
    tiers: list[TierPrice]


def _prices_out(db: Session, item: Item) -> ItemPricesOut:
    rows = db.scalars(select(ItemPrice).where(ItemPrice.item_id == item.id)).all()
    return ItemPricesOut(
        item_id=item.id, base_sale_price=item.sale_price,
        tiers=[TierPrice(tier=r.tier, price=r.price,
                         discount_pct=getattr(r, "discount_pct", 0) or 0,
                         vat_pct=getattr(r, "vat_pct", 0) or 0) for r in rows],
    )


@router.get("/{item_id}/prices", response_model=ItemPricesOut)
def get_item_prices(
    item_id: int,
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_READ)),
    db: Session = Depends(get_db),
) -> ItemPricesOut:
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    return _prices_out(db, item)


class ReturnPriceOut(BaseModel):
    item_id: int
    unit_price: Decimal
    discount_pct: Decimal
    source: str   # «last_purchase» أو «item_price» أو «none»


@router.get("/{item_id}/return-price", response_model=ReturnPriceOut)
def get_return_price(
    item_id: int,
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_READ)),
    db: Session = Depends(get_db),
) -> ReturnPriceOut:
    """السعر اللي سطر المرتجع بيتملّى بيه — آخر سعر شراء، وإلا سعر الصنف الحالي.

    المرتجع كان بيفتح بسعر فاضي، فاللي بيكتبه بيروح يدوّر على فاتورة الشراء أو يكتب رقم
    من دماغه. وآخر سعر شراء هو الرقم الصح للمرتجع: البضاعة دي دخلت بالسعر ده، فرجوعها
    بنفسه هو اللي بيخلّي المخزون والحساب يقفلوا على نفس المبلغ.

    ولو الصنف عمره ما اتشرى (اتصنّع، أو رصيد افتتاحي)، بيرجع سعر البيع الحالي — رقم ليه
    معنى أحسن من خانة فاضية، واللي مش عاجبه بيغيّره.

    الخصم بيتبع نفس الترتيب: خصم الصنف المسجّل عليه، وصفر لو مفيش.
    """
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "الصنف مش موجود"})

    from src.services import costing_service

    last = costing_service.last_purchase_cost(db, item_id)
    if last and Decimal(str(last)) > 0:
        price, source = Decimal(str(last)), "last_purchase"
    elif item.sale_price is not None and Decimal(str(item.sale_price)) > 0:
        price, source = Decimal(str(item.sale_price)), "item_price"
    else:
        price, source = Decimal("0"), "none"

    # الخصم بيتسجّل على شريحة السعر مش على الصنف — `item_price.discount_pct`. أول شريحة
    # ليها خصم هي اللي بتتاخد: الصنف اللي ليه خصم متفق عليه بيبقى نفس النسبة على شرايحه.
    row = db.scalar(
        select(ItemPrice.discount_pct)
        .where(ItemPrice.item_id == item_id, ItemPrice.discount_pct.isnot(None))
        .limit(1)
    )
    disc = Decimal(str(row or 0))

    return ReturnPriceOut(item_id=item_id, unit_price=price, discount_pct=disc, source=source)


@router.put("/{item_id}/prices", response_model=ItemPricesOut)
def set_item_prices(
    item_id: int,
    body: ItemPricesSet,
    current: CurrentUser = Depends(require_capability(CAP_CATALOG_WRITE)),
    db: Session = Depends(get_db),
) -> ItemPricesOut:
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    # Upsert each provided tier (omitted tiers are left unchanged) — the same writer the create
    # endpoint uses, so the two cannot validate or log differently.
    _apply_tiers(db, item, body.tiers, actor_user_id=current.id)
    db.commit()
    return _prices_out(db, item)


class UnitOut(BaseModel):
    name: str
    factor: Decimal
    is_base: bool


class ItemUnitsOut(BaseModel):
    item_id: int
    base_unit: str
    units: list[UnitOut]


class UnitIn(BaseModel):
    name: str
    factor: Decimal


class ItemUnitsSet(BaseModel):
    units: list[UnitIn]


def _units_out(db: Session, item: Item) -> ItemUnitsOut:
    rows = db.scalars(select(ItemUnit).where(ItemUnit.item_id == item.id)).all()
    units = [UnitOut(name=item.unit_of_measure, factor=Decimal("1.000"), is_base=True)]
    units += [UnitOut(name=r.name, factor=r.factor, is_base=False) for r in rows]
    return ItemUnitsOut(item_id=item.id, base_unit=item.unit_of_measure, units=units)


@router.get("/{item_id}/units", response_model=ItemUnitsOut)
def get_item_units(
    item_id: int,
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_READ)),
    db: Session = Depends(get_db),
) -> ItemUnitsOut:
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    return _units_out(db, item)


@router.put("/{item_id}/units", response_model=ItemUnitsOut)
def set_item_units(
    item_id: int,
    body: ItemUnitsSet,
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_WRITE)),
    db: Session = Depends(get_db),
) -> ItemUnitsOut:
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    _apply_units(db, item, body.units)
    db.commit()
    return _units_out(db, item)


class SerialOut(BaseModel):
    id: int
    item_id: int
    serial: str
    status: str
    location_kind: str | None
    location_id: int | None


class ReceiveSerials(BaseModel):
    location_kind: LocationKind
    location_id: int
    serials: list[str]


def _serial_out(s: ItemSerial) -> SerialOut:
    return SerialOut(
        id=s.id, item_id=s.item_id, serial=s.serial, status=s.status.value,
        location_kind=s.location_kind.value if s.location_kind else None,
        location_id=s.location_id,
    )


@router.get("/{item_id}/serials", response_model=list[SerialOut])
def list_serials(
    item_id: int,
    status_filter: str | None = Query(default=None, alias="status"),
    location_kind: LocationKind | None = None,
    location_id: int | None = None,
    _: CurrentUser = Depends(require_capability(CAP_STOCK_READ)),
    db: Session = Depends(get_db),
) -> list[SerialOut]:
    stmt = select(ItemSerial).where(ItemSerial.item_id == item_id)
    if status_filter is not None:
        stmt = stmt.where(ItemSerial.status == SerialStatus(status_filter))
    if location_kind is not None:
        stmt = stmt.where(ItemSerial.location_kind == location_kind)
    if location_id is not None:
        stmt = stmt.where(ItemSerial.location_id == location_id)
    return [_serial_out(s) for s in db.scalars(stmt.order_by(ItemSerial.serial)).all()]


@router.post("/{item_id}/serials/receive", response_model=list[SerialOut],
             status_code=status.HTTP_201_CREATED)
def receive_serials(
    item_id: int,
    body: ReceiveSerials,
    current: CurrentUser = Depends(require_capability(CAP_PURCHASE_WRITE)),
    db: Session = Depends(get_db),
) -> list[SerialOut]:
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    try:
        rows = serial_service.receive(
            db, item=item, location_kind=body.location_kind, location_id=body.location_id,
            serials=body.serials, actor_user_id=current.id,
        )
    except SerialError as exc:
        raise HTTPException(422, {"code": "serial_invalid", "message": str(exc)})
    db.commit()
    return [_serial_out(s) for s in rows]


class ItemProfileOut(BaseModel):
    """ملف الصنف — stock, sales, purchases, movements and price history in one call."""

    item: ItemOut
    on_hand: Decimal
    stock_by_location: list[dict] = []
    sold_quantity: Decimal
    sold_value: Decimal
    purchased_quantity: Decimal
    purchased_value: Decimal
    avg_sale_price: Decimal
    avg_purchase_price: Decimal
    sales: list[dict] = []
    purchases: list[dict] = []
    movements: list[dict] = []
    price_history: list[dict] = []
    tier_prices: list[dict] = []


@router.get("/{item_id}/balance", response_model=dict)
def item_balance(
    item_id: int,
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_READ)),
    db: Session = Depends(get_db),
) -> dict:
    """Prices + the quantity in every stock location — the stock-enquiry screen (رصيد صنف)."""
    try:
        return item_profile_service.balance(db, item_id)
    except item_profile_service.ItemProfileError as exc:
        raise HTTPException(404, {"code": "not_found", "message": str(exc)}) from exc


@router.get("/{item_id}/card", response_model=dict)
def item_card(
    item_id: int,
    location_kind: str | None = Query(None, description="warehouse | custody"),
    location_id: int | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    movement_type: str | None = Query(None),
    direction: str | None = Query(None, description="in | out"),
    _: CurrentUser = Depends(require_capability(CAP_STOCK_READ)),
    db: Session = Depends(get_db),
) -> dict:
    """كارت الصنف — every movement with the balance before and after it.

    Without a location this is the item's whole position; with one it is that store's card.
    Filters hide rows but never rewrite balances — see `src/lib/item_card.py`.
    """
    try:
        return item_card_lib.card(
            db, item_id=item_id, location_kind=location_kind, location_id=location_id,
            date_from=date_from, date_to=date_to, movement_type=movement_type,
            direction=direction,
        )
    except item_card_lib.ItemCardError as exc:
        raise HTTPException(404, {"code": "not_found", "message": str(exc)}) from exc


@router.get("/{item_id}/profile", response_model=ItemProfileOut)
def item_profile(
    item_id: int,
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_READ)),
    db: Session = Depends(get_db),
) -> ItemProfileOut:
    """Everything the system knows about one item — the product file."""
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    try:
        p = item_profile_service.profile(db, item_id)
    except item_profile_service.ItemProfileError as exc:
        raise HTTPException(404, {"code": "not_found", "message": str(exc)}) from exc
    base = _out(item)
    base.on_hand = p.on_hand
    return ItemProfileOut(
        item=base, on_hand=p.on_hand, stock_by_location=p.stock_by_location,
        sold_quantity=p.sold_quantity, sold_value=p.sold_value,
        purchased_quantity=p.purchased_quantity, purchased_value=p.purchased_value,
        avg_sale_price=p.avg_sale_price, avg_purchase_price=p.avg_purchase_price,
        sales=p.sales, purchases=p.purchases, movements=p.movements,
        price_history=p.price_history, tier_prices=p.tier_prices,
    )


@router.get("/{item_id}", response_model=ItemOut)
def get_item(
    item_id: int,
    _: CurrentUser = Depends(require_capability(CAP_CATALOG_READ)),
    db: Session = Depends(get_db),
) -> ItemOut:
    """One item's card — everything stored about it, without the 360 file's history."""
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    out = _out(item)
    out.on_hand = item_profile_service.bulk_on_hand(db, [item_id]).get(item_id, Decimal("0.000"))
    return out


@router.patch("/{item_id}", response_model=ItemOut)
def update_item(
    item_id: int,
    body: ItemUpdate,
    current: CurrentUser = Depends(require_capability(CAP_CATALOG_WRITE)),
    db: Session = Depends(get_db),
) -> ItemOut:
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    # Editing reference prices never rewrites prices already snapshotted on posted documents.
    # Price moves are logged (027) so «why did this get cheaper?» has an answer.
    PRICE_FIELDS = {"purchase_price", "sale_price", "default_discount_pct"}
    # (031) «مش مبعوت» و«مبعوت فاضي» حاجتين مختلفتين.
    #
    # This loop used to skip every None, which made a nullable field one-way: a discount, a
    # reorder level or a default warehouse could be set and then never removed. That is not a
    # small gap for `default_discount_pct` in particular — NULL means «no fixed discount on this
    # item» and 0 means «its discount is nothing», and the sale reads the two differently, so a
    # rate typed once could be changed but never withdrawn.
    #
    # Pydantic already knows which keys the caller actually sent. A field left out is untouched;
    # a field sent as null is CLEARED, but only where the column allows it — `name` and `active`
    # are not nullable, and a null there is a malformed request, not an instruction.
    # Exactly the columns that are nullable. `default_discount_pct` is deliberately NOT among
    # them: an item ALWAYS has a rate and 0 is «no discount», which is a complete answer. The
    # NULL-versus-zero distinction lives on the CUSTOMER, whose column is nullable — there, NULL
    # means «nothing agreed with him» and the sale falls back to the item's rate.
    CLEARABLE = {"purchase_price", "sale_price", "default_warehouse_id", "category",
                 "min_stock", "max_stock",
                 "piece_name", "pieces_per_unit", "description"}
    sent = body.model_fields_set
    for field in ("code", "name", "purchase_price", "sale_price", "is_serialized", "active",
                  "default_warehouse_id", "category",
                  "default_discount_pct",
                  "min_stock", "max_stock", "is_perishable",   # (011)
                  "piece_name", "pieces_per_unit", "description"):
        if field not in sent:
            continue
        val = getattr(body, field)
        if val is None and field not in CLEARABLE:
            continue
        if field in PRICE_FIELDS:
            item_profile_service.record_price_change(
                db, item_id=item.id, field_name=field, old_value=getattr(item, field),
                new_value=val, actor_user_id=current.id)
        setattr(item, field, val)
    db.flush()
    db.commit()
    return _out(item)


def _delete_item(db: Session, item: Item, actor_user_id: int) -> None:
    """Permanently remove an item that never moved — otherwise refuse.

    Deleting an item that appears on a posted invoice, a stock movement or a recipe would
    orphan those documents, so that case must stay a deactivation.
    """
    from src.models.bom import BomComponent
    from src.models.manufacturing import ManufacturingOrder
    from src.models.purchasing import PurchaseInvoiceLine
    from src.models.sales import SalesInvoiceLine
    from src.models.stock import StockMovement

    blockers: list[str] = []
    checks = [
        ("حركات مخزون", select(func.count()).select_from(StockMovement)
         .where(StockMovement.item_id == item.id)),
        ("سطور فواتير بيع", select(func.count()).select_from(SalesInvoiceLine)
         .where(SalesInvoiceLine.item_id == item.id)),
        ("سطور فواتير شراء", select(func.count()).select_from(PurchaseInvoiceLine)
         .where(PurchaseInvoiceLine.item_id == item.id)),
        ("وصفات تصنيع", select(func.count()).select_from(BomComponent)
         .where(BomComponent.item_id == item.id)),
        ("أوامر تصنيع", select(func.count()).select_from(ManufacturingOrder)
         .where(ManufacturingOrder.product_id == item.id)),
    ]
    for label, stmt in checks:
        if (db.scalar(stmt) or 0) > 0:
            blockers.append(label)

    if blockers:
        raise ValueError(
            "لا يمكن حذف الصنف نهائياً لوجود " + "، ".join(blockers)
            + ". يمكنك إلغاء تفعيله بدلاً من الحذف."
        )

    from src.models.catalog import ItemPriceHistory, ItemSerial, ItemUnit
    from src.models.loyalty import ProductPointValue

    audit_service.record(db, action="item.delete", actor_user_id=actor_user_id,
                         entity_type="item", entity_id=item.id,
                         before={"code": item.code, "name": item.name})
    # Owned rows carry no history of their own once the item is gone.
    for model in (ItemPrice, ItemUnit, ItemSerial, ItemPriceHistory,
                  ProductPointValue):
        db.execute(delete(model).where(model.item_id == item.id))
    db.delete(item)
    db.flush()


@router.delete("/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def deactivate_item(
    item_id: int,
    hard: bool = False,
    current: CurrentUser = Depends(require_capability(CAP_CATALOG_WRITE)),
    db: Session = Depends(get_db),
) -> None:
    """Deactivate the item; `hard=true` deletes it outright — only if it never moved."""
    item = db.get(Item, item_id)
    if item is None:
        raise HTTPException(404, {"code": "not_found", "message": "Item not found"})
    if hard:
        try:
            _delete_item(db, item, current.id)
        except ValueError as exc:
            raise HTTPException(409, {"code": "has_history", "message": str(exc)}) from exc
        db.commit()
        return
    item.active = False  # soft-delete: never hard-delete an item referenced by posted documents
    db.flush()
    db.commit()
